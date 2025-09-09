import os
import pandas as pd
from flask import Flask, jsonify
from fpdf import FPDF
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from openpyxl.styles import PatternFill
from flask import render_template

app = Flask(__name__)

# ---------- Sample Student Data ----------
students = [
    {"RollNo": "101", "Name": "Raj", "Attendance": 60, "Marks": 40, "FeesPaid": False},
    {"RollNo": "102", "Name": "Neha", "Attendance": 85, "Marks": 20, "FeesPaid": True},
    {"RollNo": "103", "Name": "Aman", "Attendance": 92, "Marks": 75, "FeesPaid": True},
]

# ---------- Student Dashboard ----------
@app.route("/", methods=["GET"])
def student_dashboard():
    student_data = []

    for s in students:
        reasons = []
        if s["Attendance"] < 75:
            reasons.append("Low Attendance")
        if s["Marks"] < 35:
            reasons.append("Low Marks")
        if not s["FeesPaid"]:
            reasons.append("Unpaid Fees")

        if len(reasons) >= 2:
            level = "danger"
            label = "High Risk"
        elif len(reasons) == 1:
            level = "warning"
            label = "Warning"
        else:
            level = "safe"
            label = "Safe"

        student_data.append({**s, "risk_level": level, "risk_label": label})

    return render_template("students.html", students=student_data)


# Ensure reports folder exists
REPORT_DIR = "reports"
os.makedirs(REPORT_DIR, exist_ok=True)

# ---------- Risk Detection ----------
def detect_risks(students):
    risked = []
    for s in students:
        reason = []
        if s["Attendance"] < 75:
            reason.append("Low Attendance")
        if s["Marks"] < 35:
            reason.append("Low Marks")
        if not s["FeesPaid"]:
            reason.append("Unpaid Fees")
        if reason:
            risked.append({
                "RollNo": s["RollNo"],
                "Name": s["Name"],
                "RiskReason": ", ".join(reason)
            })
    return risked

# ---------- Excel Generation ----------
def generate_excel(risk_students):
    df = pd.DataFrame(risk_students)
    file_path = os.path.join(REPORT_DIR, "at_risk_students.xlsx")

    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Risked")
        workbook = writer.book
        sheet = writer.sheets["Risked"]

        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        for row in range(2, len(df) + 2):
            for col in range(1, len(df.columns) + 1):
                sheet.cell(row=row, column=col).fill = red_fill

    return file_path

# ---------- PDF Report ----------
def generate_pdf(risk_students):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    pdf.cell(200, 10, txt="At-Risk Students Report", ln=True, align="C")

    for s in risk_students:
        pdf.cell(200, 10, txt=f"{s['RollNo']} - {s['Name']} - {s['RiskReason']}", ln=True)

    file_path = os.path.join(REPORT_DIR, "at_risk_students.pdf")
    pdf.output(file_path)
    return file_path

# ---------- Email Sender ----------
def send_email_with_attachment(file_path):
    sender_email = "aayadav371@gmail.com"
    password = "mnfm ufeu decl hxex"  # 🔐 Use an App Password here
    # receiver_email = "aayadav973@gmail.com"
    receiver_email = "mp7702524@gmail.com"

    msg = MIMEMultipart()
    msg["From"] = sender_email
    msg["To"] = receiver_email
    msg["Subject"] = "At-Risk Students Report"

    body = "Dear Mentor,\n\nPlease find attached the latest At-Risk Students report."
    msg.attach(MIMEText(body, "plain"))

    try:
        with open(file_path, "rb") as f:
            attach = MIMEApplication(f.read(), _subtype=file_path.split(".")[-1])
            attach.add_header("Content-Disposition", "attachment", filename=os.path.basename(file_path))
            msg.attach(attach)

        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()
        server.login(sender_email, password)
        server.send_message(msg)
        server.quit()

        print(f"[INFO] Email sent successfully to {receiver_email} with attachment: {file_path}")
    except Exception as e:
        print(f"[ERROR] Failed to send email: {e}")

# ---------- Flask Route ----------
@app.route("/get_risk_students", methods=["GET"])
def get_risk_students():
    risked = detect_risks(students)

    # Generate Excel and PDF reports
    excel_file = generate_excel(risked)
    pdf_file = generate_pdf(risked)

    # Send to mentor
    send_email_with_attachment(excel_file)
    send_email_with_attachment(pdf_file)

    # Return JSON for frontend
    return jsonify(risked)

# ---------- Run Server ----------
if __name__ == "__main__":
    app.run(debug=True)
